# -*- coding: utf-8 -*-
from __future__ import unicode_literals

from django.shortcuts import render
from io import BytesIO
from openpyxl import load_workbook

from .forms import FileUploadForm
from .models import FileContents

# Create your views here.
def fileUploadView(request):
	status = ""
	x=''
	if request.method == "POST":
		form = FileUploadForm(request.POST, request.FILES)
		if form.is_valid():
			file = request.FILES["xlsFile"].read()
			wb = load_workbook(filename=BytesIO(file))
			ws = wb.active
			for row in ws.iter_rows(min_row=1, max_row=3, max_col=2):
				(r1,r2) = row 
				dbrow = FileContents.objects.create(row1=r1.value, row2=r2.value)
				dbrow.save()
			form.save()
			status = "Success!" + ws['A1'].value + x
	else:
		form = FileUploadForm()
	extractedContents = FileContents.objects.all()
	return render(request, "home.html", {'form' : form, 'status' : status, 'extractedContents' : extractedContents})